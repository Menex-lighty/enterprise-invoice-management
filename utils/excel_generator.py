"""
Excel Generator
Handles Excel file generation for reports and data exports
"""

import os
from datetime import datetime
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, LineChart, PieChart

class ExcelGenerator:
    """Excel generation utility class"""
    
    def __init__(self, output_dir='generated_files'):
        self.output_dir = output_dir
        self.ensure_output_dir()
        
    def ensure_output_dir(self):
        """Ensure output directory exists"""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
    
    def generate_invoice_excel(self, invoice, company=None, customer=None, filename=None):
        """Generate Excel file for a single invoice"""
        try:
            if not filename:
                filename = f"invoice_{invoice.invoice_number.replace('/', '_')}.xlsx"
            
            filepath = os.path.join(self.output_dir, filename)
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Invoice"
            
            # Styling
            header_font = Font(bold=True, size=14)
            title_font = Font(bold=True, size=16)
            table_header_font = Font(bold=True, size=12)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Company header
            row = 1
            if company:
                ws.cell(row=row, column=1, value=company.name).font = title_font
                row += 1
                
                if company.address:
                    ws.cell(row=row, column=1, value=company.address)
                    row += 1
                
                if company.city and company.state:
                    ws.cell(row=row, column=1, value=f"{company.city}, {company.state}")
                    row += 1
                
                if company.pincode:
                    ws.cell(row=row, column=1, value=f"PIN: {company.pincode}")
                    row += 1
                
                if company.gstin:
                    ws.cell(row=row, column=1, value=f"GSTIN: {company.gstin}")
                    row += 1
                
                if company.contact_phone:
                    ws.cell(row=row, column=1, value=f"Phone: {company.contact_phone}")
                    row += 1
                
                if company.email:
                    ws.cell(row=row, column=1, value=f"Email: {company.email}")
                    row += 1
                
                row += 1
            
            # Invoice title
            ws.cell(row=row, column=1, value="TAX INVOICE").font = header_font
            row += 2
            
            # Invoice details
            ws.cell(row=row, column=1, value="Invoice Number:").font = header_font
            ws.cell(row=row, column=2, value=invoice.invoice_number)
            row += 1
            
            ws.cell(row=row, column=1, value="Invoice Date:").font = header_font
            ws.cell(row=row, column=2, value=invoice.invoice_date.strftime('%d-%m-%Y') if invoice.invoice_date else '')
            row += 1
            
            if invoice.po_number:
                ws.cell(row=row, column=1, value="PO Number:").font = header_font
                ws.cell(row=row, column=2, value=invoice.po_number)
                row += 1
            
            if invoice.po_date:
                ws.cell(row=row, column=1, value="PO Date:").font = header_font
                ws.cell(row=row, column=2, value=invoice.po_date.strftime('%d-%m-%Y'))
                row += 1
            
            if invoice.payment_mode:
                ws.cell(row=row, column=1, value="Payment Mode:").font = header_font
                ws.cell(row=row, column=2, value=invoice.payment_mode)
                row += 1
            
            row += 1
            
            # Customer details
            if customer:
                ws.cell(row=row, column=1, value="Bill To:").font = header_font
                row += 1
                
                ws.cell(row=row, column=1, value="Customer:").font = header_font
                ws.cell(row=row, column=2, value=customer.name)
                row += 1
                
                if customer.address:
                    ws.cell(row=row, column=1, value="Address:").font = header_font
                    ws.cell(row=row, column=2, value=customer.address)
                    row += 1
                
                if customer.city and customer.state:
                    ws.cell(row=row, column=1, value="City, State:").font = header_font
                    ws.cell(row=row, column=2, value=f"{customer.city}, {customer.state}")
                    row += 1
                
                if customer.pincode:
                    ws.cell(row=row, column=1, value="PIN Code:").font = header_font
                    ws.cell(row=row, column=2, value=customer.pincode)
                    row += 1
                
                if customer.gstin:
                    ws.cell(row=row, column=1, value="GSTIN:").font = header_font
                    ws.cell(row=row, column=2, value=customer.gstin)
                    row += 1
                
                if customer.contact_person:
                    ws.cell(row=row, column=1, value="Contact Person:").font = header_font
                    ws.cell(row=row, column=2, value=customer.contact_person)
                    row += 1
                
                if customer.phone:
                    ws.cell(row=row, column=1, value="Phone:").font = header_font
                    ws.cell(row=row, column=2, value=customer.phone)
                    row += 1
                
                row += 1
            
            # Items table
            if invoice.items:
                # Headers
                headers = ['S.No.', 'Description', 'HSN Code', 'Quantity', 'Unit', 'Rate', 'Discount %', 'Amount']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = table_header_font
                    cell.border = border
                    cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                
                row += 1
                
                # Items
                for i, item in enumerate(invoice.items, 1):
                    hsn_code = item.product.hsn_code if item.product and item.product.hsn_code else ''
                    
                    data = [
                        i,
                        item.description or '',
                        hsn_code,
                        float(item.quantity) if item.quantity else 0,
                        item.unit or '',
                        float(item.rate) if item.rate else 0,
                        float(item.discount_percent) if item.discount_percent else 0,
                        float(item.amount) if item.amount else 0
                    ]
                    
                    for col, value in enumerate(data, 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.border = border
                        if col >= 4:  # Numeric columns
                            cell.alignment = Alignment(horizontal='right')
                    
                    row += 1
                
                # Totals
                totals_row = row
                ws.cell(row=totals_row, column=7, value="Subtotal:").font = table_header_font
                ws.cell(row=totals_row, column=8, value=float(invoice.subtotal) if invoice.subtotal else 0)
                
                row += 1
                ws.cell(row=row, column=7, value="GST (18%):").font = table_header_font
                ws.cell(row=row, column=8, value=float(invoice.gst_amount) if invoice.gst_amount else 0)
                
                row += 1
                ws.cell(row=row, column=7, value="Total:").font = table_header_font
                ws.cell(row=row, column=8, value=float(invoice.total_amount) if invoice.total_amount else 0)
                
                # Style totals
                for r in range(totals_row, row + 1):
                    for c in range(7, 9):
                        cell = ws.cell(row=r, column=c)
                        cell.border = border
                        cell.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
                        if c == 8:
                            cell.alignment = Alignment(horizontal='right')
                
                row += 2
            
            # Banking details
            if company and company.bank_name:
                ws.cell(row=row, column=1, value="Banking Details:").font = header_font
                row += 1
                
                ws.cell(row=row, column=1, value="Bank Name:")
                ws.cell(row=row, column=2, value=company.bank_name)
                row += 1
                
                if company.account_number:
                    ws.cell(row=row, column=1, value="Account Number:")
                    ws.cell(row=row, column=2, value=company.account_number)
                    row += 1
                
                if company.ifsc_code:
                    ws.cell(row=row, column=1, value="IFSC Code:")
                    ws.cell(row=row, column=2, value=company.ifsc_code)
                    row += 1
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 15
            
            # Save workbook
            wb.save(filepath)
            
            return filepath
            
        except Exception as e:
            raise Exception(f"Error generating Excel file: {str(e)}")
    
    def generate_invoices_report(self, invoices, filename=None):
        """Generate Excel report for multiple invoices"""
        try:
            if not filename:
                filename = f"invoices_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            filepath = os.path.join(self.output_dir, filename)
            
            # Prepare data
            data = []
            for invoice in invoices:
                data.append({
                    'Invoice Number': invoice.invoice_number,
                    'Invoice Date': invoice.invoice_date.strftime('%d-%m-%Y') if invoice.invoice_date else '',
                    'Customer': invoice.customer.name if invoice.customer else '',
                    'PO Number': invoice.po_number or '',
                    'PO Date': invoice.po_date.strftime('%d-%m-%Y') if invoice.po_date else '',
                    'Subtotal': float(invoice.subtotal) if invoice.subtotal else 0,
                    'GST Amount': float(invoice.gst_amount) if invoice.gst_amount else 0,
                    'Total Amount': float(invoice.total_amount) if invoice.total_amount else 0,
                    'Status': invoice.status,
                    'Payment Mode': invoice.payment_mode or '',
                    'Created Date': invoice.created_at.strftime('%d-%m-%Y') if invoice.created_at else ''
                })
            
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Invoices Report"
            
            # Add title
            ws.cell(row=1, column=1, value="INVOICES REPORT").font = Font(bold=True, size=16)
            ws.cell(row=2, column=1, value=f"Generated on: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}")
            
            # Add data
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Style header
            header_row = 4
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=header_row, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                   top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Auto-adjust column widths
            for column in ws.columns:
                length = max(len(str(cell.value)) for cell in column)
                ws.column_dimensions[column[0].column_letter].width = min(length + 2, 50)
            
            # Add summary sheet
            ws_summary = wb.create_sheet("Summary")
            
            # Summary calculations
            total_invoices = len(invoices)
            total_amount = sum(float(inv.total_amount) if inv.total_amount else 0 for inv in invoices)
            draft_count = sum(1 for inv in invoices if inv.status == 'DRAFT')
            sent_count = sum(1 for inv in invoices if inv.status == 'SENT')
            paid_count = sum(1 for inv in invoices if inv.status == 'PAID')
            cancelled_count = sum(1 for inv in invoices if inv.status == 'CANCELLED')
            
            # Add summary data
            summary_data = [
                ['Invoice Summary', ''],
                ['Total Invoices', total_invoices],
                ['Total Amount', total_amount],
                ['', ''],
                ['Status Breakdown', ''],
                ['Draft', draft_count],
                ['Sent', sent_count],
                ['Paid', paid_count],
                ['Cancelled', cancelled_count]
            ]
            
            for row_num, (label, value) in enumerate(summary_data, 1):
                ws_summary.cell(row=row_num, column=1, value=label).font = Font(bold=True)
                ws_summary.cell(row=row_num, column=2, value=value)
            
            # Save workbook
            wb.save(filepath)
            
            return filepath
            
        except Exception as e:
            raise Exception(f"Error generating invoices report: {str(e)}")
    
    def generate_customers_report(self, customers, filename=None):
        """Generate Excel report for customers"""
        try:
            if not filename:
                filename = f"customers_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            filepath = os.path.join(self.output_dir, filename)
            
            # Prepare data
            data = []
            for customer in customers:
                data.append({
                    'Customer Name': customer.name,
                    'Address': customer.address or '',
                    'City': customer.city or '',
                    'State': customer.state or '',
                    'PIN Code': customer.pincode or '',
                    'GSTIN': customer.gstin or '',
                    'Contact Person': customer.contact_person or '',
                    'Phone': customer.phone or '',
                    'Email': customer.email or '',
                    'Created Date': customer.created_at.strftime('%d-%m-%Y') if customer.created_at else '',
                    'Invoice Count': len(customer.invoices)
                })
            
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # Create workbook and save
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Customers', index=False)
                
                # Get workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Customers']
                
                # Style header
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    length = max(len(str(cell.value)) for cell in column)
                    worksheet.column_dimensions[column[0].column_letter].width = min(length + 2, 50)
            
            return filepath
            
        except Exception as e:
            raise Exception(f"Error generating customers report: {str(e)}")
    
    def generate_products_report(self, products, filename=None):
        """Generate Excel report for products"""
        try:
            if not filename:
                filename = f"products_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            filepath = os.path.join(self.output_dir, filename)
            
            # Prepare data
            data = []
            for product in products:
                data.append({
                    'Category': product.category or '',
                    'Product Name': product.name,
                    'Description': product.description or '',
                    'Unit': product.unit,
                    'Rate': float(product.rate) if product.rate else 0,
                    'HSN Code': product.hsn_code or '',
                    'Created Date': product.created_at.strftime('%d-%m-%Y') if product.created_at else '',
                    'Usage Count': len(product.invoice_items)
                })
            
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Products Report"
            
            # Add title
            ws.cell(row=1, column=1, value="PRODUCTS REPORT").font = Font(bold=True, size=16)
            ws.cell(row=2, column=1, value=f"Generated on: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}")
            
            # Add data
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Style header
            header_row = 4
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=header_row, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                   top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Auto-adjust column widths
            for column in ws.columns:
                length = max(len(str(cell.value)) for cell in column)
                ws.column_dimensions[column[0].column_letter].width = min(length + 2, 50)
            
            # Add category analysis sheet
            ws_categories = wb.create_sheet("Categories")
            
            # Group by category
            category_counts = {}
            category_avg_rates = {}
            
            for product in products:
                category = product.category or 'Uncategorized'
                if category not in category_counts:
                    category_counts[category] = 0
                    category_avg_rates[category] = []
                
                category_counts[category] += 1
                if product.rate:
                    category_avg_rates[category].append(float(product.rate))
            
            # Calculate averages
            for category in category_avg_rates:
                if category_avg_rates[category]:
                    category_avg_rates[category] = sum(category_avg_rates[category]) / len(category_avg_rates[category])
                else:
                    category_avg_rates[category] = 0
            
            # Add category analysis
            ws_categories.cell(row=1, column=1, value="CATEGORY ANALYSIS").font = Font(bold=True, size=14)
            ws_categories.cell(row=3, column=1, value="Category").font = Font(bold=True)
            ws_categories.cell(row=3, column=2, value="Product Count").font = Font(bold=True)
            ws_categories.cell(row=3, column=3, value="Average Rate").font = Font(bold=True)
            
            row = 4
            for category, count in category_counts.items():
                ws_categories.cell(row=row, column=1, value=category)
                ws_categories.cell(row=row, column=2, value=count)
                ws_categories.cell(row=row, column=3, value=round(category_avg_rates[category], 2))
                row += 1
            
            # Save workbook
            wb.save(filepath)
            
            return filepath
            
        except Exception as e:
            raise Exception(f"Error generating products report: {str(e)}")
    
    def generate_excel_buffer(self, data, headers, sheet_name="Data"):
        """Generate Excel file in memory and return as BytesIO buffer"""
        try:
            buffer = BytesIO()
            
            # Create DataFrame
            df = pd.DataFrame(data, columns=headers)
            
            # Write to buffer
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Get workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                # Style header
                for col in range(1, len(headers) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    length = max(len(str(cell.value)) for cell in column)
                    worksheet.column_dimensions[column[0].column_letter].width = min(length + 2, 50)
            
            buffer.seek(0)
            return buffer
            
        except Exception as e:
            raise Exception(f"Error generating Excel buffer: {str(e)}")